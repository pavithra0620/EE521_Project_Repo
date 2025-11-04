import os
import copy
import streamlit as st
import pandas as pd
import numpy as np
import networkx as nx
import plotly.graph_objects as go
import urllib.parse
try:
    from streamlit_plotly_events import plotly_events
    HAS_PLOTLY_EVENTS = True
except Exception:
    HAS_PLOTLY_EVENTS = False
import re
import io
import openpyxl
from openpyxl.utils import get_column_letter
from matpower_parser import load_case
from powerflow import build_ybus, newton_raphson, get_bus_types
from matpower_adapter import run_matpower_pf, matlab_engine_available
from pypower_adapter import run_pypower_pf, pypower_available
from n1 import run_n1_sweep_serial, run_n1_sweep_parallel

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

st.set_page_config(layout='wide', page_title='Contingency Analysis Dashboard')
st.title('Contingency Analysis — Line Outage')

# Helpers

def compute_branch_flows(case, V):
    """Compute branch complex power from both ends for each branch.

    Uses a simple pi-model approximate formula with tap ratio handling:
      I_from = (V_from / a - V_to) * y
      S_from = V_from * conj(I_from)
      I_to   = (V_to - V_from / a) * y
      S_to   = V_to * conj(I_to)

    Returns list of dicts with keys: f, t, P_from, Q_from, S_from, P_to, Q_to, S_to, rateA, status
    """
    flows = []
    nb = len(case['bus'])
    for br in case.get('branch', []):
        # ensure branch has enough entries
        if len(br) < 4:
            continue
        f = int(br[0]) - 1
        t = int(br[1]) - 1
        r = float(br[2])
        x = float(br[3])
        b = float(br[4]) if len(br) > 4 else 0.0
        rateA = float(br[5]) if len(br) > 5 else 0.0
        ratio = float(br[8]) if len(br) > 8 else 0.0
        angle = float(br[9]) if len(br) > 9 else 0.0
        status = int(br[10]) if len(br) > 10 else 1
        if status == 0:
            # treat as out-of-service
            flows.append({
                'f': f+1, 't': t+1, 'P_from': 0.0, 'Q_from': 0.0, 'S_from': 0.0,
                'P_to': 0.0, 'Q_to': 0.0, 'S_to': 0.0, 'rateA': rateA, 'status': status
            })
            continue
        z = r + 1j * x
        if abs(z) < 1e-12:
            y = 0+0j
        else:
            y = 1.0 / z
        a = ratio if (ratio not in (0.0, None)) else 1.0
        # handle indexing safety
        try:
            Vf = V[f]
            Vt = V[t]
        except Exception:
            Vf = 1.0+0j
            Vt = 1.0+0j
        # approximate branch currents and powers
        # I_from = (Vf / a - Vt) * y
        # I_to   = (Vt - Vf / a) * y
        # use complex tap angle if angle provided
        if a == 0:
            a = 1.0
        I_from = (Vf / a - Vt) * y
        I_to = (Vt - Vf / a) * y
        S_from = Vf * np.conj(I_from)
        S_to = Vt * np.conj(I_to)
        flows.append({
            'f': f+1, 't': t+1,
            'P_from': float(np.real(S_from)), 'Q_from': float(np.imag(S_from)), 'S_from': abs(S_from),
            'P_to': float(np.real(S_to)), 'Q_to': float(np.imag(S_to)), 'S_to': abs(S_to),
            'rateA': rateA, 'status': status
        })
    return flows


def run_pf_and_flows(case, prefer_matpower: bool = False):
    # If prefer_matpower is True, try MATLAB/MATPOWER first.
    # Otherwise prefer PyPower, then MATPOWER, then Python NR solver.
    if prefer_matpower and matlab_engine_available():
        try:
            mpres = run_matpower_pf(case)
            V = mpres.get('V')
            if V is not None:
                res = {'V': V}
                flows = compute_branch_flows(case, V)
                return res, flows
        except Exception as e:
            import streamlit as _st
            _st.warning(f'MATPOWER run failed: {e}; falling back to other solvers')

    # Prefer PyPower if available, then MATPOWER via MATLAB, otherwise Python NR solver.
    if pypower_available():
        try:
            pr = run_pypower_pf(case)
            V = pr.get('V')
            if V is not None:
                res = {'V': V}
                flows = compute_branch_flows(case, V)
                return res, flows
        except Exception as e:
            import streamlit as _st
            _st.warning(f'PyPower run failed: {e}; trying MATPOWER or Python solver')

    if matlab_engine_available():
        try:
            mpres = run_matpower_pf(case)
            V = mpres.get('V')
            # if MATPOWER returned only raw struct or V is None, fall back to Python solver
            if V is None:
                raise RuntimeError('MATPOWER returned no voltages; falling back to Python solver')
            res = {'V': V}
            flows = compute_branch_flows(case, V)
            return res, flows
        except Exception as e:
            import streamlit as _st
            _st.warning(f'MATPOWER run failed: {e}; falling back to Python solver')

    # run Python power flow fallback
    res = newton_raphson(case, tol=1e-8, maxiter=200, damping=0.5)
    V = res['V']
    flows = compute_branch_flows(case, V)
    return res, flows

# UI: load case
st.sidebar.header('Case selection')
example = st.sidebar.selectbox('Example case', ['case9', 'case5 (small)'])
if example == 'case9':
    case = load_case(os.path.join(BASE_DIR, 'cases', 'case9.json'))
else:
    # create a minimal 5-bus demo if requested
    case = {
        'baseMVA': 100,
        'bus': [[1,3,0,0,0,0,1,1.06,0,230,1,1.06,0.94],
                [2,1,100,60,0,0,1,1.045,0,230,1,1.06,0.94],
                [3,1,90,40,0,0,1,1.01,0,230,1,1.06,0.94],
                [4,1,120,80,0,0,1,1.02,0,230,1,1.06,0.94],
                [5,1,60,30,0,0,1,1.01,0,230,1,1.06,0.94]],
        'gen': [[1, 500, 0, 9999, -9999, 1.06, 100, 1, 500, 0]],
        'branch': [[1,2,0.02,0.06,0.03,250,0,0,0,0,1,-360,360],
                   [1,3,0.08,0.24,0.025,250,0,0,0,0,1,-360,360],
                   [2,3,0.06,0.18,0.02,150,0,0,0,0,1,-360,360],
                   [2,4,0.06,0.18,0.02,150,0,0,0,0,1,-360,360],
                   [3,5,0.04,0.12,0.01,150,0,0,0,0,1,-360,360]]
    }

st.sidebar.write(f"Loaded case: baseMVA={case.get('baseMVA')} buses={len(case['bus'])} branches={len(case['branch'])}")

# UI: color customization and interaction options
healthy_color = st.sidebar.color_picker('Healthy line color', '#2ca02c')
overloaded_color = st.sidebar.color_picker('Overloaded line color', '#d62728')
removed_color = st.sidebar.color_picker('Removed line color', '#636363')
bg_opacity = st.sidebar.slider('Background opacity', min_value=0.0, max_value=1.0, value=0.25, step=0.05)
auto_apply_on_click = st.sidebar.checkbox('Apply branch removal immediately when clicked', value=True)

# Legend / color preview
st.sidebar.markdown('**Legend**')
st.sidebar.markdown(f"<div style='display:flex;gap:8px;align-items:center'><div style='width:18px;height:12px;background:{healthy_color};border:1px solid #ccc'></div> Healthy &nbsp; <div style='width:18px;height:12px;background:{overloaded_color};border:1px solid #ccc'></div> Overloaded &nbsp; <div style='width:18px;height:12px;background:{removed_color};border:1px solid #ccc'></div> Removed</div>", unsafe_allow_html=True)

# session state for selected branch
if 'selected_branch' not in st.session_state:
    st.session_state['selected_branch'] = 0

# run base case PF
with st.spinner('Running base-case power flow...'):
    base_res, base_flows = run_pf_and_flows(case)

st.subheader('Base case results')
col1, col2 = st.columns([2,1])
with col1:
    V = base_res['V']
    df_bus = pd.DataFrame([{'bus': i+1, 'V_mag': float(abs(V[i])), 'V_ang_deg': float(np.angle(V[i], deg=True))} for i in range(len(V))])
    # presentable column names for display only (keep original df_bus for computations)
    df_bus_display = df_bus.rename(columns={
        'V_mag': 'Voltage (p.u.)',
        'V_ang_deg': 'Angle (deg)'
    })
    st.table(df_bus_display.set_index('bus'))
with col2:
    st.write('Branch flows (from-end S [MVA])')
    df_flows = pd.DataFrame(base_flows)
    # compute loading percent relative to rateA for base flows
    df_flows['S_from_MVA'] = df_flows['S_from']
    def _loading_pct(row):
        try:
            rate = float(row.get('rateA', 0) or 0.0)
            s = float(row.get('S_from', 0.0) or 0.0)
            if rate > 0:
                return 100.0 * s / rate
        except Exception:
            pass
        return None
    df_flows['loading_pct'] = df_flows.apply(_loading_pct, axis=1)
    # friendly column names (no underscores)
    df_display = df_flows[['f','t','P_from','Q_from','S_from','rateA','loading_pct']].rename(columns={
        'f': 'From', 't': 'To',
        'P_from': 'P from (MW)', 'Q_from': 'Q from (MVAr)', 'S_from': 'S from (MVA)',
        'rateA': 'Rate A (MVA)', 'loading_pct': 'Loading %'
    }).round(3)
    st.dataframe(df_display)

st.markdown('---')
st.subheader('Contingency selection')
# show interactive network and selectable list of branches
G = nx.Graph()
for b in case['bus']:
    G.add_node(int(b[0]))
for i, br in enumerate(case['branch']):
    try:
        status = int(br[10]) if len(br) > 10 else 1
    except Exception:
        status = 1
    if status == 0:
        continue
    G.add_edge(int(br[0]), int(br[1]), index=i)
pos = nx.spring_layout(G, seed=42)

# Geolocation mock options
use_geo = st.sidebar.checkbox('Use geolocated layout (mock)', value=False)
# default to Pullman, Washington for the mock map
geo_center_lat = st.sidebar.number_input('Map center lat', value=46.7324, format="%.6f")
geo_center_lon = st.sidebar.number_input('Map center lon', value=-117.1654, format="%.6f")
geo_spread_km = st.sidebar.number_input('Mock spread (km)', min_value=0.1, max_value=50.0, value=5.0)
# optional CSV upload for real bus coordinates (columns: bus, lat, lon)
coords_upload = st.sidebar.file_uploader('Upload bus lat/lon CSV (columns: bus, lat, lon)', type=['csv'])
# example coords helper
example_coords_path = os.path.join(BASE_DIR, 'data', 'example_coords_pullman.csv')
if st.sidebar.button('Load example coords (Pullman, WA)'):
    try:
        coords_df_example = pd.read_csv(example_coords_path)
        # write to a small in-memory buffer so later code can read from coords_upload-like object
        coords_upload = io.BytesIO()
        coords_df_example.to_csv(coords_upload, index=False)
        coords_upload.seek(0)
        st.sidebar.success('Loaded example coords into map.')
    except Exception as e:
        st.sidebar.error(f'Failed to load example coords: {e}')

# build plotly network with colored edges and a lightweight SVG city map background
def _make_city_svg(width=800, height=600):
    # simple stylized city map: grid of blocks and roads
    svg_parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}">',
        '<rect width="100%" height="100%" fill="#f8f8f8"/>',
    ]
    # draw roads (horizontal and vertical)
    for i in range(1,6):
        y = int(i * height / 6)
        svg_parts.append(f'<rect x="0" y="{y-8}" width="{width}" height="16" fill="#e0e0e0"/>')
    for j in range(1,6):
        x = int(j * width / 6)
        svg_parts.append(f'<rect x="{x-8}" y="0" width="16" height="{height}" fill="#e0e0e0"/>')
    # add some green parks
    svg_parts.append(f'<rect x="{width*0.05}" y="{height*0.05}" width="{width*0.15}" height="{height*0.12}" fill="#d9f0d3"/>')
    svg_parts.append(f'<rect x="{width*0.7}" y="{height*0.6}" width="{width*0.18}" height="{height*0.18}" fill="#d9f0d3"/>')
    svg_parts.append('</svg>')
    svg = '\n'.join(svg_parts)
    return 'data:image/svg+xml;utf8,' + urllib.parse.quote(svg)

city_img = _make_city_svg()

# determine overloaded branches in base case (S_from > rateA)
overloaded_base = set()
for i, f in enumerate(base_flows):
    rate = f.get('rateA', 0) or 0.0
    s = f.get('S_from', 0.0) or 0.0
    if rate > 0 and s > rate:
        overloaded_base.add(i+1)

node_x = []
node_y = []
node_text = []
for n in G.nodes():
    x,y = pos[n]
    node_x.append(x)
    node_y.append(y)
    node_text.append(f'Bus {n}')

edge_traces = []
for u,v,d in G.edges(data=True):
    bi = int(d.get('index', 0)) + 1
    x0, y0 = pos[u]
    x1, y1 = pos[v]
    color = healthy_color if bi not in overloaded_base else overloaded_color
    # include S, rate, and loading percent in hover text for clarity
    hover_s = ''
    try:
        bf = base_flows[bi-1]
        s_val = float(bf.get('S_from', 0.0) or 0.0)
        rate_val = float(bf.get('rateA', 0.0) or 0.0)
        loading_val = (100.0 * s_val / rate_val) if rate_val > 0 else None
        hover_s = f"<br>S from: {s_val:.2f} MVA<br>Rate A: {rate_val:.1f} MVA"
        if loading_val is not None:
            hover_s += f"<br>Loading: {loading_val:.1f}%"
    except Exception:
        hover_s = ''
    edge_traces.append(go.Scatter(x=[x0, x1], y=[y0, y1], mode='lines', line=dict(width=4, color=color), hoverinfo='text', text=f'Branch {bi}: {u}-{v}' + hover_s))

node_trace = go.Scatter(x=node_x, y=node_y, mode='markers+text', text=[str(n) for n in G.nodes()], textposition='top center', marker=dict(size=20, color='#7b2cbf'))

fig = go.Figure(data=edge_traces + [node_trace])
fig.update_layout(
    showlegend=False,
    hovermode='closest',
    margin=dict(b=20,l=5,r=5,t=40),
    xaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
    yaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
    images=[dict(source=city_img, xref='paper', yref='paper', x=0, y=1, sizex=1, sizey=1, sizing='stretch', opacity=bg_opacity, layer='below')]
)
if not use_geo:
    st.plotly_chart(fig, use_container_width=True)
    # capture click on schematic if the optional component is installed
    if HAS_PLOTLY_EVENTS:
        ev = plotly_events(fig, click_event=True, hover_event=False)
        if ev:
            txt = ev[0].get('text','')
            m = re.search(r"(\d+)", txt or '')
            if m:
                idx = int(m.group(1)) - 1
                st.session_state['selected_branch'] = idx
                if auto_apply_on_click:
                    st.session_state['do_contingency'] = True
else:
    # create mock lat/lon positions from spring layout, centered at provided geo center
    # convert spread km to degrees roughly (1 deg ~ 111 km)
    deg_spread = float(geo_spread_km) / 111.0
    xs = np.array(node_x)
    ys = np.array(node_y)
    # normalize to [-0.5,0.5]
    xsn = (xs - xs.mean()) / (np.ptp(xs) if np.ptp(xs) != 0 else 1.0)
    ysn = (ys - ys.mean()) / (np.ptp(ys) if np.ptp(ys) != 0 else 1.0)
    lats = geo_center_lat + ysn * deg_spread
    lons = geo_center_lon + xsn * deg_spread

    # if the user uploaded a CSV of bus coordinates, override mock positions where available
    coords_map = {}
    if coords_upload is not None:
        try:
            df_coords = pd.read_csv(coords_upload)
            # expect columns 'bus', 'lat', 'lon' (bus numbers as ints)
            for _, r in df_coords.iterrows():
                try:
                    b = int(r['bus'])
                    coords_map[b] = (float(r['lat']), float(r['lon']))
                except Exception:
                    continue
        except Exception:
            st.sidebar.warning('Could not read uploaded coords CSV — expected columns: bus, lat, lon')

    # build mapbox traces for edges
    edge_map_traces = []
    for tr in edge_traces:
        # each tr has x,y points in layout coords; map to node indices by matching text
        txt = tr.text if hasattr(tr, 'text') else None
        if txt and isinstance(txt, str) and txt.startswith('Branch'):
            parts = txt.split(':',1)[0].split()
            try:
                bidx = int(parts[1]) - 1
            except Exception:
                bidx = None
        else:
            bidx = None
        # fallback: use positions by iterating edges from G
        # We'll iterate edges of G in the same insertion order to map them
        pass

    # instead, create edges directly from G with lat/lon
    map_edges = []
    for u,v,d in G.edges(data=True):
        ui = list(G.nodes()).index(u)
        vi = list(G.nodes()).index(v)
        # allow uploaded coords to override mock positions
        if coords_upload is not None and (u in coords_map or v in coords_map):
            try:
                la, lo = coords_map.get(u, (lats[ui], lons[ui]))
                lb, lob = coords_map.get(v, (lats[vi], lons[vi]))
            except Exception:
                la, lo = lats[ui], lons[ui]
                lb, lob = lats[vi], lons[vi]
        else:
            la = lats[ui]; lo = lons[ui]; lb = lats[vi]; lob = lons[vi]
        idx = int(d.get('index', 0))
        branch_no = idx + 1
        # include hover details
        try:
            bf = base_flows[idx]
            s_val = float(bf.get('S_from', 0.0) or 0.0)
            rate_val = float(bf.get('rateA', 0.0) or 0.0)
            loading_val = (100.0 * s_val / rate_val) if rate_val > 0 else None
            hover_s = f"S from: {s_val:.2f} MVA | Rate A: {rate_val:.1f} MVA"
            if loading_val is not None:
                hover_s += f" | Loading: {loading_val:.1f}%"
        except Exception:
            hover_s = ''
        color = healthy_color if branch_no not in overloaded_base else overloaded_color
        edge_map_traces.append(go.Scattermapbox(lat=[la, lb], lon=[lo, lob], mode='lines', line=dict(width=4,color=color), hoverinfo='text', text=f'Branch {branch_no}: {u}-{v} - {hover_s}'))

    node_map = go.Scattermapbox(lat=lats, lon=lons, mode='markers+text', marker=dict(size=14,color='#7b2cbf'), text=[str(n) for n in G.nodes()], textposition='top right')
    map_fig = go.Figure(data=edge_map_traces + [node_map])
    map_fig.update_layout(mapbox_style='open-street-map', mapbox=dict(center=dict(lat=geo_center_lat, lon=geo_center_lon), zoom=10), margin=dict(b=20,l=5,r=5,t=40), showlegend=False)
    st.plotly_chart(map_fig, use_container_width=True)
    if HAS_PLOTLY_EVENTS:
        ev = plotly_events(map_fig, click_event=True, hover_event=False)
        if ev:
            txt = ev[0].get('text','')
            m = re.search(r"(\d+)", txt or '')
            if m:
                idx = int(m.group(1)) - 1
                st.session_state['selected_branch'] = idx
                if auto_apply_on_click:
                    st.session_state['do_contingency'] = True

# branch selector
branch_labels = [f"{i+1}: {int(br[0])}-{int(br[1])} (rate {br[5] if len(br)>5 else 'n/a'})" for i,br in enumerate(case['branch'])]
sel = st.selectbox('Select branch to remove (contingency)', options=list(range(len(case['branch']))), index=st.session_state.get('selected_branch', 0), format_func=lambda i: branch_labels[i])
st.session_state['selected_branch'] = sel

if st.button('Run contingency'):
    # don't call show_contingency here because the function is defined later in the file
    # instead set the session flags so the deferred call (below) will run it after
    st.session_state['selected_branch'] = sel
    st.session_state['do_contingency'] = True

# N-1 sweep: run all single-line outages and rank by severity
def run_n1_sweep(case, base_V, base_flows):
    results = []
    nb = len(case.get('branch', []))
    for i in range(nb):
        br = case['branch'][i]
        # skip already out-of-service branches
        try:
            status = int(br[10]) if len(br) > 10 else 1
        except Exception:
            status = 1
        if status == 0:
            continue
        case2 = copy.deepcopy(case)
        # set branch status to 0
        if len(case2['branch'][i]) > 10:
            case2['branch'][i][10] = 0
        else:
            while len(case2['branch'][i]) <= 10:
                case2['branch'][i].append(0)
            case2['branch'][i][10] = 0
        try:
            res2, flows2 = run_pf_and_flows(case2)
            V2 = res2.get('V')
            success = V2 is not None
        except Exception as e:
            flows2 = []
            V2 = None
            success = False

        # compute metrics
        max_loading = float('nan')
        overload_count = 0
        if flows2:
            ratios = []
            for f in flows2:
                rate = f.get('rateA', 0) or 0.0
                s = f.get('S_from', 0.0) or 0.0
                if rate > 0:
                    ratios.append(s / rate)
            if ratios:
                max_loading = max(ratios)
                overload_count = sum(1 for r in ratios if r > 1.0)

        max_v_drop = float('nan')
        if base_V is not None and V2 is not None:
            try:
                mags_base = np.abs(base_V)
                mags_cont = np.abs(V2)
                # align lengths
                L = min(len(mags_base), len(mags_cont))
                max_v_drop = float(np.max(mags_base[:L] - mags_cont[:L]))
            except Exception:
                max_v_drop = float('nan')

        results.append({
            'branch': i+1,
            'from': int(br[0]),
            'to': int(br[1]),
            'success': bool(success),
            'overload_count': int(overload_count),
            'max_loading_ratio': float(max_loading) if not np.isnan(max_loading) else None,
            'max_voltage_drop': float(max_v_drop) if not np.isnan(max_v_drop) else None,
        })
    return results

# N-1 sweep controls
st.markdown('---')
st.subheader('N-1 sweep')
col_run, col_opts = st.columns([2,1])
with col_opts:
    parallel = st.checkbox('Parallel sweep (multi-process)', value=False)
    max_workers = None
    if parallel:
        max_workers = st.number_input('Max workers', min_value=1, max_value=16, value=4)
    auto_run = st.checkbox('Auto-run N-1 on load', value=False)
    subset = st.multiselect('Subset of branches to sweep (leave empty = all)', options=branch_labels)
    # weights for severity score
    st.markdown('Severity score weights')
    w_over = st.number_input('Weight: overload_count', value=1000.0)
    w_load = st.number_input('Weight: max_loading_ratio_excess', value=100.0)
    w_vdrop = st.number_input('Weight: max_voltage_drop', value=10.0)

def _parse_subset(subset_list):
    if not subset_list:
        return None
    idxs = []
    for s in subset_list:
        # format: '1: from-to (rate ...)'
        try:
            i = int(s.split(':',1)[0]) - 1
            idxs.append(i)
        except Exception:
            pass
    return idxs

def _run_sweep_and_report(case, base_res, base_flows, parallel=False, max_workers=None, subset=None, weights=(1000.0,100.0,10.0)):
    base_V = base_res.get('V')
    idxs = _parse_subset(subset)
    if parallel:
        # run parallel and show progress
        with st.spinner('Running parallel N-1 sweep...'):
            results = run_n1_sweep_parallel(case, base_V=base_V, branches=idxs, max_workers=int(max_workers) if max_workers else None)
    else:
        with st.spinner('Running serial N-1 sweep...'):
            results = run_n1_sweep_serial(case, base_V=base_V, branches=idxs)
    # apply weights (recompute severity with chosen weights)
    for r in results:
        # recompute score
        lc = r.get('overload_count', 0)
        mlr = r.get('max_loading_ratio')
        mvd = r.get('max_voltage_drop')
        # default computation inline to avoid import
        load_excess = 0.0
        if mlr is not None:
            load_excess = max(0.0, mlr - 1.0)
        vdrop = max(0.0, mvd or 0.0)
        score = lc * weights[0] + load_excess * weights[1] + vdrop * weights[2]
        r['severity_score'] = float(score)
    df_sweep = pd.DataFrame(results)
    df_sweep = df_sweep.sort_values(by=['severity_score'], ascending=False)
    st.subheader('N-1 contingency ranking')
    st.dataframe(df_sweep[['branch','from','to','success','overload_count','max_loading_ratio','max_voltage_drop','severity_score']].reset_index(drop=True))
    csv = df_sweep.to_csv(index=False)
    st.download_button('Download N-1 results (CSV)', csv, file_name='n1_results.csv')


def show_contingency(sel_idx: int):
    """Run the contingency for branch index sel_idx and render results/visuals."""
    case2 = copy.deepcopy(case)
    # set branch status to 0 for the selected branch
    if len(case2['branch'][sel_idx]) > 10:
        case2['branch'][sel_idx][10] = 0
    else:
        while len(case2['branch'][sel_idx]) <= 10:
            case2['branch'][sel_idx].append(0)
        case2['branch'][sel_idx][10] = 0

    with st.spinner('Running contingency power flow (MATPOWER)...'):
        # force MATPOWER (MATLAB) run if available for most accurate comparison
        res2, flows2 = run_pf_and_flows(case2, prefer_matpower=True)

    st.subheader('Post-contingency results')
    colA, colB = st.columns([2,1])
    with colA:
        V2 = res2['V']
        df_bus2 = pd.DataFrame([{'bus': i+1, 'V_mag': float(abs(V2[i])), 'V_ang_deg': float(np.angle(V2[i], deg=True))} for i in range(len(V2))])
        # display-friendly copy
        df_bus2_display = df_bus2.rename(columns={
            'V_mag': 'Voltage (p.u.)',
            'V_ang_deg': 'Angle (deg)'
        })
        st.table(df_bus2_display.set_index('bus'))
    with colB:
        df_flows2 = pd.DataFrame(flows2)
        # compute loading percent for contingency flows as well
        def _loading_pct_row(row):
            try:
                rate = float(row.get('rateA', 0) or 0.0)
                s = float(row.get('S_from', 0.0) or 0.0)
                if rate > 0:
                    return 100.0 * s / rate
            except Exception:
                pass
            return None

        df_flows2['loading_pct'] = df_flows2.apply(_loading_pct_row, axis=1)
        df2_display = df_flows2[['f','t','P_from','Q_from','S_from','rateA','loading_pct']].rename(columns={
            'f': 'From', 't': 'To',
            'P_from': 'P from (MW)', 'Q_from': 'Q from (MVAr)', 'S_from': 'S from (MVA)',
            'rateA': 'Rate A (MVA)', 'loading_pct': 'Loading %'
        }).round(3)
        st.dataframe(df2_display)

    # compare and flag overloads
    st.markdown('---')
    st.subheader('Contingency comparison')
    comp_bus = df_bus.merge(df_bus2, on='bus', suffixes=('_base','_cont'))
    comp_bus['V_drop'] = comp_bus['V_mag_base'] - comp_bus['V_mag_cont']
    st.write('Bus voltages comparison (base vs contingency)')
    comp_bus = comp_bus.rename(columns={
        'V_mag_base': 'Voltage (base) (p.u.)',
        'V_mag_cont': 'Voltage (cont) (p.u.)',
        'V_drop': 'Voltage drop (p.u.)'
    })
    # Presentable styling: highlight big drops (>0.05 p.u.) and low voltages (<0.95 p.u.)
    comp_bus_display = comp_bus.set_index('bus').round(6)
    def _highlight_voltage(row):
        styles = []
        try:
            vdrop = float(row.get('Voltage drop (p.u.)') or 0.0)
        except Exception:
            vdrop = 0.0
        try:
            vcont = float(row.get('Voltage (cont) (p.u.)') or 0.0)
        except Exception:
            vcont = 1.0
        for col in comp_bus_display.columns:
            if col == 'Voltage drop (p.u.)' and vdrop > 0.05:
                styles.append('background-color:#fff2cc')
            elif col == 'Voltage (cont) (p.u.)' and vcont < 0.95:
                styles.append('background-color:#ffdddd')
            else:
                styles.append('')
        return styles
    try:
        st.dataframe(comp_bus_display.style.apply(_highlight_voltage, axis=1))
    except Exception:
        st.dataframe(comp_bus_display)

    # branch flow comparison
    df_base_flows = pd.DataFrame(base_flows)
    df_cont_flows = pd.DataFrame(flows2)
    df_base_flows['branch'] = df_base_flows.index + 1
    df_cont_flows['branch'] = df_cont_flows.index + 1
    comp_flow = df_base_flows[['branch','f','t','S_from','rateA']].merge(df_cont_flows[['branch','S_from']], on='branch', suffixes=('_base','_cont'))
    # treat non-positive or missing rateA as no limit (not overloaded)
    def _is_overloaded(row):
        try:
            rate = float(row.get('rateA', 0) or 0.0)
            s_cont = float(row.get('S_from_cont', 0.0) or 0.0)
            if rate <= 0:
                return False
            return s_cont > rate
        except Exception:
            return False
    comp_flow['overload'] = comp_flow.apply(_is_overloaded, axis=1)
    # compute loading percentages and percent change relative to base
    def _safe_pct(s, rate):
        try:
            rate = float(rate or 0.0)
            s = float(s or 0.0)
            if rate > 0:
                return 100.0 * s / rate
        except Exception:
            pass
        return None

    comp_flow['loading_pct_base'] = comp_flow.apply(lambda r: _safe_pct(r['S_from_base'], r['rateA']), axis=1)
    comp_flow['loading_pct_cont'] = comp_flow.apply(lambda r: _safe_pct(r['S_from_cont'], r['rateA']), axis=1)
    # change in loading percent (contingency - base)
    comp_flow['loading_pct_change'] = comp_flow.apply(lambda r: (r['loading_pct_cont'] - r['loading_pct_base']) if (r['loading_pct_cont'] is not None and r['loading_pct_base'] is not None) else None, axis=1)
    # percent over capacity (how much percent above 100% in contingency)
    comp_flow['percent_over_capacity'] = comp_flow['loading_pct_cont'].apply(lambda x: (x - 100.0) if (x is not None and x > 100.0) else 0.0)

    st.write('Branch flow comparison (from-end MVA):')
    comp_display = comp_flow[['branch','f','t','S_from_base','S_from_cont','rateA','loading_pct_base','loading_pct_cont','loading_pct_change','percent_over_capacity','overload']].rename(columns={
        'branch': 'Branch', 'f': 'From', 't': 'To',
        'S_from_base': 'S from (base) (MVA)', 'S_from_cont': 'S from (cont) (MVA)',
        'rateA': 'Rate A (MVA)', 'loading_pct_base': 'Loading % (base)', 'loading_pct_cont': 'Loading % (cont)',
        'loading_pct_change': 'Loading % change', 'percent_over_capacity': 'Percent over capacity', 'overload': 'Overload'
    }).round(3)
    # present boolean as Yes/No for overload
    if 'Overload' in comp_display.columns:
        comp_display['Overload'] = comp_display['Overload'].apply(lambda x: 'Yes' if bool(x) else 'No')
    # summary metrics
    try:
        max_over = float(comp_display['Percent over capacity'].max()) if len(comp_display) > 0 else 0.0
    except Exception:
        max_over = 0.0
    try:
        num_over = int((comp_display['Overload'] == 'Yes').sum()) if 'Overload' in comp_display.columns else 0
    except Exception:
        num_over = 0
    c1, c2 = st.columns([1,1])
    c1.metric('Max % over capacity', f"{max_over:.1f}%")
    c2.metric('Branches overloaded', str(num_over))

    # allow CSV / Excel download of the comparison table (presentable columns)
    try:
        csv_bytes = comp_display.to_csv(index=False).encode('utf-8')
        st.download_button('Download comparison (CSV)', data=csv_bytes, file_name='contingency_comparison.csv', mime='text/csv')
    except Exception:
        pass

    # Excel export (optional, requires openpyxl)
    try:
        # create an Excel file with formatting where possible
        out = io.BytesIO()
        with pd.ExcelWriter(out, engine='openpyxl') as writer:
            comp_display.to_excel(writer, index=False, sheet_name='comparison')
            writer.book  # ensure book exists
            writer.save()
        out.seek(0)
        # post-process with openpyxl to set column widths and number formats
        try:
            wb = openpyxl.load_workbook(out)
            ws = wb['comparison']
            for i, col in enumerate(comp_display.columns, start=1):
                col_letter = get_column_letter(i)
                # width heuristic
                ws.column_dimensions[col_letter].width = min(max(12, len(col) + 2), 40)
                # number format heuristics
                if 'Loading' in col or 'Percent' in col:
                    fmt = '0.0'
                elif 'Voltage' in col:
                    fmt = '0.000'
                elif 'S from' in col or 'Rate A' in col:
                    fmt = '0.000'
                else:
                    fmt = None
                if fmt:
                    for cell in ws[col_letter]:
                        # skip header cell
                        if cell.row == 1:
                            continue
                        cell.number_format = fmt
            out2 = io.BytesIO()
            wb.save(out2)
            out2.seek(0)
            st.download_button('Download comparison (Excel)', data=out2, file_name='contingency_comparison.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        except Exception:
            # fallback to plain bytes if post-processing fails
            out.seek(0)
            st.download_button('Download comparison (Excel)', data=out, file_name='contingency_comparison.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception:
        st.sidebar.info('Excel export requires pandas & openpyxl; install them to enable Excel downloads')

    # highlight overloaded rows (red background) for easier scanning
    def _highlight_overload(row):
        try:
            return ['background-color:#ffdddd' if row.get('Overload') == 'Yes' else '' for _ in row.index]
        except Exception:
            return ['' for _ in row.index]

    try:
        st.dataframe(comp_display.style.apply(_highlight_overload, axis=1))
    except Exception:
        # fallback to plain table if styler fails
        st.dataframe(comp_display)

    # build visualization: color removed branch and overloaded
    overloaded = set(comp_flow[comp_flow['overload']]['branch'].tolist())
    removed = sel_idx + 1
    nodes_list = list(G.nodes())
    edge_traces_local = []
    for i, br in enumerate(case['branch']):
        try:
            u = int(br[0]); v = int(br[1])
        except Exception:
            continue
        try:
            ui = nodes_list.index(u); vi = nodes_list.index(v)
        except ValueError:
            continue
        if not use_geo:
            x0,y0 = pos[u]; x1,y1 = pos[v]
            color = removed_color if (i+1) == removed else (overloaded_color if (i+1) in overloaded else healthy_color)
            dash = 'dash' if (i+1) == removed else 'solid'
            edge_traces_local.append(go.Scatter(x=[x0,x1], y=[y0,y1], mode='lines', line=dict(width=4,color=color,dash=dash), hoverinfo='text', text=f'Branch {i+1}: {u}-{v}'))
        else:
            la = lats[ui]; lo = lons[ui]; lb = lats[vi]; lob = lons[vi]
            color = removed_color if (i+1) == removed else (overloaded_color if (i+1) in overloaded else healthy_color)
            edge_traces_local.append(go.Scattermapbox(lat=[la,lb], lon=[lo,lob], mode='lines', line=dict(width=4,color=color), hoverinfo='text', text=f'Branch {i+1}: {u}-{v}'))

    if not use_geo:
        node_trace2 = go.Scatter(x=node_x, y=node_y, mode='markers+text', text=[str(n) for n in G.nodes()], textposition='top center', marker=dict(size=20, color='#7b2cbf'))
        fig2 = go.Figure(data=edge_traces_local + [node_trace2])
        fig2.update_layout(showlegend=False, hovermode='closest', margin=dict(b=20,l=5,r=5,t=40), xaxis=dict(showgrid=False, zeroline=False, showticklabels=False), yaxis=dict(showgrid=False, zeroline=False, showticklabels=False), images=[dict(source=city_img, xref='paper', yref='paper', x=0, y=1, sizex=1, sizey=1, sizing='stretch', opacity=bg_opacity, layer='below')])
        st.plotly_chart(fig2, use_container_width=True)
    else:
        node_idxs = [nodes_list.index(n) for n in G.nodes()]
        node_lats = [lats[i] for i in node_idxs]
        node_lons = [lons[i] for i in node_idxs]
        node_map2 = go.Scattermapbox(lat=node_lats, lon=node_lons, mode='markers+text', marker=dict(size=14,color='#7b2cbf'), text=[str(n) for n in G.nodes()], textposition='top right')
        map_fig2 = go.Figure(data=edge_traces_local + [node_map2])
        map_fig2.update_layout(mapbox_style='open-street-map', mapbox=dict(center=dict(lat=geo_center_lat, lon=geo_center_lon), zoom=10), margin=dict(b=20,l=5,r=5,t=40), showlegend=False)
        st.plotly_chart(map_fig2, use_container_width=True)

# If a click set the do_contingency flag, run it now
if st.session_state.get('do_contingency'):
    st.session_state['do_contingency'] = False
    sel_now = st.session_state.get('selected_branch', 0)
    show_contingency(sel_now)

if auto_run:
    _run_sweep_and_report(case, base_res, base_flows, parallel=parallel, max_workers=max_workers, subset=subset, weights=(w_over,w_load,w_vdrop))

if st.button('Run N-1 sweep (all single-line contingencies)'):
    _run_sweep_and_report(case, base_res, base_flows, parallel=parallel, max_workers=max_workers, subset=subset, weights=(w_over,w_load,w_vdrop))


st.markdown('---')
st.info('This contingency tool uses the Python NR solver included in this package (a MATPOWER-like approach). For strict MATPOWER results you would run MATLAB/Matpower externally and import results.')
